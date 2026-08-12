from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass, fields
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Mapping, Sequence


@dataclass(frozen=True)
class Product:
    source_row: int
    name: str
    sku: str
    price: str
    source_type: str = "simple"
    parent: str = ""
    compare_at_price: str = ""
    description: str = ""
    brand: str = ""
    category: str = ""
    stock: str = "0"
    image_url: str = ""
    source_url: str = ""
    barcode: str = ""
    weight_kg: str = ""
    color: str = ""
    size: str = ""
    handle: str = ""


@dataclass(frozen=True)
class Issue:
    source_row: int
    sku: str
    name: str
    code: str
    message: str


ALIASES = {
    "source_type": ("type", "product type"),
    "name": ("name", "product name", "title", "product title", "item name"),
    "sku": ("sku", "product sku", "variant sku", "article", "article number"),
    "price": ("price", "regular price", "selling price", "sale price"),
    "compare_at_price": ("compare at price", "compare_at_price", "msrp", "rrp"),
    "description": ("description", "product description", "body", "body html", "body (html)"),
    "brand": ("brand", "vendor", "manufacturer"),
    "category": ("category", "categories", "product category", "collection"),
    "stock": ("stock", "quantity", "inventory", "inventory quantity", "qty"),
    "image_url": ("image", "image url", "image src", "images", "photo", "photo url"),
    "source_url": ("source", "source url", "supplier url", "product url", "url"),
    "barcode": ("barcode", "ean", "upc", "gtin"),
    "weight_kg": ("weight", "weight kg", "weight (kg)"),
    "color": ("color", "colour"),
    "size": ("size",),
    "handle": ("handle", "slug"),
    "parent": ("parent", "parent sku"),
}

NORMALIZED_HEADERS = [field.name for field in fields(Product)]
ISSUE_HEADERS = [field.name for field in fields(Issue)]

SHOPIFY_HEADERS = [
    "Handle",
    "Title",
    "Body (HTML)",
    "Vendor",
    "Product Category",
    "Type",
    "Tags",
    "Published",
    "Option1 Name",
    "Option1 Value",
    "Option2 Name",
    "Option2 Value",
    "Variant SKU",
    "Variant Inventory Tracker",
    "Variant Inventory Qty",
    "Variant Inventory Policy",
    "Variant Fulfillment Service",
    "Variant Price",
    "Variant Compare At Price",
    "Variant Requires Shipping",
    "Variant Taxable",
    "Variant Barcode",
    "Image Src",
    "Image Position",
    "Image Alt Text",
    "SEO Title",
    "SEO Description",
    "Status",
]

WOOCOMMERCE_HEADERS = [
    "Type",
    "SKU",
    "Name",
    "Published",
    "Is featured?",
    "Visibility in catalog",
    "Short description",
    "Description",
    "Tax status",
    "In stock?",
    "Stock",
    "Backorders allowed?",
    "Sold individually?",
    "Weight (kg)",
    "Allow customer reviews?",
    "Sale price",
    "Regular price",
    "Categories",
    "Tags",
    "Images",
    "External URL",
    "Position",
    "Parent",
    "Attribute 1 name",
    "Attribute 1 value(s)",
    "Attribute 1 visible",
    "Attribute 1 global",
    "Attribute 2 name",
    "Attribute 2 value(s)",
    "Attribute 2 visible",
    "Attribute 2 global",
]


def _header_key(value: object) -> str:
    text = str(value or "").strip().lower().replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", text)


def _clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load_csv(path: Path) -> tuple[list[str], list[list[object]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            header = next(reader)
        except StopIteration:
            return [], []
        return [str(cell) for cell in header], [list(row) for row in reader]


def _load_xlsx(path: Path) -> tuple[list[str], list[list[object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency error is actionable
        raise RuntimeError("XLSX support requires openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return [], []
    return [_clean(cell) for cell in header], [list(row) for row in rows]


def _load_rows(path: Path) -> tuple[list[str], list[list[object]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(path)
    raise ValueError(f"Unsupported input format: {suffix or '<none>'}. Use CSV or XLSX.")


def _map_headers(headers: Sequence[str]) -> dict[str, int]:
    indexed = {_header_key(header): index for index, header in enumerate(headers)}
    mapped: dict[str, int] = {}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            if alias in indexed:
                mapped[canonical] = indexed[alias]
                break
    return mapped


def _value(row: Sequence[object], mapping: Mapping[str, int], field: str) -> str:
    index = mapping.get(field)
    if index is None or index >= len(row):
        return ""
    return _clean(row[index])


def _money(value: str) -> str:
    if not value:
        return ""
    cleaned = re.sub(r"[^0-9,.-]", "", value).replace(" ", "")
    if cleaned.count(",") == 1 and "." not in cleaned:
        cleaned = cleaned.replace(",", ".")
    else:
        cleaned = cleaned.replace(",", "")
    try:
        amount = Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(value) from exc
    if amount < 0:
        raise ValueError(value)
    return f"{amount:.2f}"


def _stock(value: str) -> str:
    if not value:
        return "0"
    try:
        amount = int(Decimal(value.replace(",", ".")))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(value) from exc
    if amount < 0:
        raise ValueError(value)
    return str(amount)


def _slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "product"


def _write_csv(path: Path, headers: Sequence[str], rows: Iterable[Mapping[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _product_dict(product: Product) -> dict[str, object]:
    return {field.name: getattr(product, field.name) for field in fields(Product)}


def _issue_dict(issue: Issue) -> dict[str, object]:
    return {field.name: getattr(issue, field.name) for field in fields(Issue)}


def _shopify_row(product: Product) -> dict[str, str]:
    return {
        "Handle": product.handle,
        "Title": product.name,
        "Body (HTML)": product.description,
        "Vendor": product.brand,
        "Product Category": product.category,
        "Type": product.category.split(">")[-1].strip() if product.category else "",
        "Tags": ", ".join(value for value in (product.brand, product.color, product.size) if value),
        "Published": "FALSE",
        "Option1 Name": "Color" if product.color else ("Size" if product.size else "Title"),
        "Option1 Value": product.color or product.size or "Default Title",
        "Option2 Name": "Size" if product.color and product.size else "",
        "Option2 Value": product.size if product.color and product.size else "",
        "Variant SKU": product.sku,
        "Variant Inventory Tracker": "shopify",
        "Variant Inventory Qty": product.stock,
        "Variant Inventory Policy": "deny",
        "Variant Fulfillment Service": "manual",
        "Variant Price": product.price,
        "Variant Compare At Price": product.compare_at_price,
        "Variant Requires Shipping": "TRUE",
        "Variant Taxable": "TRUE",
        "Variant Barcode": product.barcode,
        "Image Src": product.image_url,
        "Image Position": "1" if product.image_url else "",
        "Image Alt Text": product.name if product.image_url else "",
        "SEO Title": product.name,
        "SEO Description": re.sub(r"<[^>]+>", "", product.description)[:160],
        "Status": "draft",
    }


def _woocommerce_row(product: Product) -> dict[str, str]:
    return {
        "Type": product.source_type,
        "SKU": product.sku,
        "Name": product.name,
        "Published": "0",
        "Is featured?": "0",
        "Visibility in catalog": "visible",
        "Short description": re.sub(r"<[^>]+>", "", product.description)[:240],
        "Description": product.description,
        "Tax status": "taxable",
        "In stock?": "1" if int(product.stock) > 0 else "0",
        "Stock": product.stock,
        "Backorders allowed?": "0",
        "Sold individually?": "0",
        "Weight (kg)": product.weight_kg,
        "Allow customer reviews?": "1",
        "Sale price": product.price if product.compare_at_price else "",
        "Regular price": product.compare_at_price or product.price,
        "Categories": product.category,
        "Tags": ", ".join(value for value in (product.brand, product.color, product.size) if value),
        "Images": product.image_url,
        "External URL": product.source_url,
        "Position": "0",
        "Parent": product.parent,
        "Attribute 1 name": "Color" if product.color else ("Size" if product.size else ""),
        "Attribute 1 value(s)": product.color or product.size,
        "Attribute 1 visible": "1" if product.color or product.size else "",
        "Attribute 1 global": "0" if product.color or product.size else "",
        "Attribute 2 name": "Size" if product.color and product.size else "",
        "Attribute 2 value(s)": product.size if product.color and product.size else "",
        "Attribute 2 visible": "1" if product.color and product.size else "",
        "Attribute 2 global": "0" if product.color and product.size else "",
    }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def prepare_catalog(
    input_path: str | Path,
    output_dir: str | Path,
    *,
    stores: Sequence[str] = ("shopify", "woocommerce"),
    limit: int | None = None,
    job_id: str = "",
    revenue_usd: str | Decimal | None = None,
    expense_usd: str | Decimal | None = None,
) -> dict[str, object]:
    """Prepare a supplier catalog and return a receipt for the completed batch."""
    source = Path(input_path).expanduser().resolve()
    destination = Path(output_dir).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    unknown_stores = set(stores) - {"shopify", "woocommerce"}
    if unknown_stores:
        raise ValueError(f"Unsupported store targets: {', '.join(sorted(unknown_stores))}")
    if limit is not None and limit < 1:
        raise ValueError("limit must be at least 1")
    revenue = _money(str(revenue_usd)) if revenue_usd is not None else ""
    expense = _money(str(expense_usd)) if expense_usd is not None else ""

    headers, rows = _load_rows(source)
    if limit is not None:
        rows = rows[:limit]
    mapping = _map_headers(headers)
    required_columns = {"name", "sku", "price"}
    missing_columns = required_columns - set(mapping)
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing_columns))}")

    products: list[Product] = []
    issues: list[Issue] = []
    seen_skus: set[str] = set()

    for source_row, row in enumerate(rows, start=2):
        name = _value(row, mapping, "name")
        sku = _value(row, mapping, "sku")
        source_type = _value(row, mapping, "source_type").lower() or "simple"
        normalized_sku = sku.casefold()
        row_issues: list[Issue] = []

        if not name:
            row_issues.append(Issue(source_row, sku, name, "missing_name", "Product name is required"))
        if not sku:
            row_issues.append(Issue(source_row, sku, name, "missing_sku", "SKU is required"))
        elif normalized_sku in seen_skus:
            row_issues.append(Issue(source_row, sku, name, "duplicate_sku", "SKU already appeared in this batch"))

        try:
            price = _money(_value(row, mapping, "price"))
            if not price and source_type != "variable":
                raise ValueError("")
        except ValueError:
            price = ""
            row_issues.append(Issue(source_row, sku, name, "invalid_price", "Price must be a non-negative number"))

        try:
            compare_at_price = _money(_value(row, mapping, "compare_at_price"))
        except ValueError:
            compare_at_price = ""
            row_issues.append(
                Issue(source_row, sku, name, "invalid_compare_at_price", "Compare-at price must be a non-negative number")
            )

        try:
            stock = _stock(_value(row, mapping, "stock"))
        except ValueError:
            stock = "0"
            row_issues.append(Issue(source_row, sku, name, "invalid_stock", "Stock must be a non-negative integer"))

        if sku and normalized_sku not in seen_skus:
            seen_skus.add(normalized_sku)
        if row_issues:
            issues.extend(row_issues)
            continue

        handle = _value(row, mapping, "handle") or _slugify(f"{name}-{sku}")
        products.append(
            Product(
                source_row=source_row,
                name=name,
                sku=sku,
                price=price,
                source_type=source_type,
                parent=_value(row, mapping, "parent"),
                compare_at_price=compare_at_price,
                description=_value(row, mapping, "description"),
                brand=_value(row, mapping, "brand"),
                category=_value(row, mapping, "category"),
                stock=stock,
                image_url=_value(row, mapping, "image_url"),
                source_url=_value(row, mapping, "source_url"),
                barcode=_value(row, mapping, "barcode"),
                weight_kg=_value(row, mapping, "weight_kg"),
                color=_value(row, mapping, "color"),
                size=_value(row, mapping, "size"),
                handle=handle,
            )
        )

    destination.mkdir(parents=True, exist_ok=True)
    output_files: list[Path] = []

    normalized_path = destination / "products.normalized.csv"
    _write_csv(normalized_path, NORMALIZED_HEADERS, (_product_dict(product) for product in products))
    output_files.append(normalized_path)

    issues_path = destination / "issues.csv"
    _write_csv(issues_path, ISSUE_HEADERS, (_issue_dict(issue) for issue in issues))
    output_files.append(issues_path)

    if "shopify" in stores:
        shopify_path = destination / "shopify_import.csv"
        _write_csv(shopify_path, SHOPIFY_HEADERS, (_shopify_row(product) for product in products))
        output_files.append(shopify_path)
    if "woocommerce" in stores:
        woocommerce_path = destination / "woocommerce_import.csv"
        _write_csv(woocommerce_path, WOOCOMMERCE_HEADERS, (_woocommerce_row(product) for product in products))
        output_files.append(woocommerce_path)

    issue_counts = Counter(issue.code for issue in issues)
    receipt: dict[str, object] = {
        "agent": "eoh-catalog-agent",
        "version": "0.1.1",
        "job_id": job_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "input_file": str(source),
        "input_sha256": _sha256(source),
        "input_rows": len(rows),
        "ready_products": len(products),
        "blocked_products": len({issue.source_row for issue in issues}),
        "duplicate_skus": issue_counts["duplicate_sku"],
        "issue_counts": dict(sorted(issue_counts.items())),
        "targets": list(stores),
        "publish_mode": "draft",
        "output_files": [str(path) for path in output_files],
    }
    if revenue or expense:
        revenue_amount = Decimal(revenue or "0")
        expense_amount = Decimal(expense or "0")
        receipt["economics"] = {
            "revenue_usd": f"{revenue_amount:.2f}",
            "expense_usd": f"{expense_amount:.2f}",
            "net_profit_usd": f"{revenue_amount - expense_amount:.2f}",
        }
    receipt_path = destination / "receipt.json"
    receipt_path.write_text(json.dumps(receipt, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return receipt
